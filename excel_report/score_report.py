import MySQLdb
from openpyxl import Workbook, load_workbook
from openpyxl.chart.area_chart import AreaChart
from openpyxl.chart.reference import Reference


class ScoreReport(object):
    def __init__(self):
        self.wb = load_workbook('./static/tmpl.xlsx')
        self.ws = self.wb.active
        self.ws.title = '成绩统计'
        self.ws.sheet_properties.tabColor = 'ff0000'

    def get_conn(self):
        """获取mysql链接"""
        try:
            conn = MySQLdb.connect(
            db='user_grade',
            host='localhost',
            user='root',
            password='',
            charset='utf8',
            )
        except:
            pass
        return conn

    def export_data(self):
        """导出数据库数据到excel"""
        conn = self.get_conn()
        cursor = conn.cursor()
        #查询语句
        sql = 'SELECT `year`, `max`, `avg` FROM `score`'
        cursor.execute(sql)
        rows = cursor.fetchall()

        #写入数据
        row_id = 10
        for (i, row) in enumerate(rows):
            (self.ws['C{0}'.format(row_id)],
            self.ws['D{0}'.format(row_id)],
            self.ws['E{0}'.format(row_id)]) = row
            row_id += 1

        # 显示图表
        chart = AreaChart()
        chart.title = "得分统计表"
        chart.style = 13
        chart.x_axis.title = '年分'
        chart.y_axis.title = '分数'

        # 横坐标
        cats = Reference(self.ws, min_col=3, min_row=10, max_row=row_id)
        # 数据,min_row包含标题行
        data = Reference(self.ws, min_col=4, min_row=9, max_col=5, max_row=row_id)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        self.ws.add_chart(chart, 'A{0}'.format(row_id + 3))

        # 保存excel
        self.wb.save('./static/score_report.xlsx')

if __name__ == "__main__":
    client = ScoreReport()
    client.export_data()
