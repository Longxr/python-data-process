
from datetime import datetime

import MySQLdb
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, colors


class ExcelUtils(object):
    """
    Excel文件工具类
    """
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws_two = self.wb.create_sheet('我的表单')
        self.ws_three = self.wb.create_sheet()

    def do_sth(self):
        self.ws['A1'] = 66
        self.ws['A2'] = 'Hello'
        self.ws['A3'] = datetime.now()

        for row in self.ws_two['A1:E5']:
            for cell in row:
                cell.value = 2

        # 求和
        self.ws_two['G1'] = '=SUM(A1:E1)'

        # 插入图片
        img = Image('./static/test.jpg')
        self.ws.add_image(img, 'K25')

        # 合并单元格
        self.ws.merge_cells('A4:E5')
        self.ws.unmerge_cells('A4:E5')

        # 设置文字
        font = Font(sz=18, color=colors.RED)
        self.ws['A2'].font = font
        self.wb.save('./static/test.xlsx')

    def read_xls(self):
        """
        读取excel文件
        """
        ws = load_workbook('./static/score.xlsx')
        names = ws.get_sheet_names()
        print(names)
        # wb = ws.active
        wb = ws[names[0]]
        
        for (i, row) in enumerate(wb.rows):
            if i < 2:
                continue
            year = wb['A{0}'.format(i + 1)].value
            ma = wb['B{0}'.format(i + 1)].value
            avg = wb['C{0}'.format(i + 1)].value

            if year is None:
                continue

            conn = self.get_conn()
            cursor = conn.cursor()
            sql = 'INSERT INTO `score` (`year`, `max`, `avg`) VALUES({year}, {max}, {avg})'.format(year=year, max=ma, avg=avg)
            print(sql)
            cursor.execute(sql)
            conn.autocommit(True)

    def get_conn(self):
        """获取mysql链接"""
        try:
            conn = MySQLdb.connect(
            db='user_grade',
            host='localhost',
            user='root',
            password='',
            charset='utf8',
            )
        except:
            pass
        return conn

    def export_xls(self):
        """导出数据库数据到excel"""
        conn = self.get_conn()
        cursor = conn.cursor()
        #查询语句
        sql = 'SELECT `year`, `max`, `avg` FROM `score`'
        #数据量大则分页查询
        # sql = 'SELECT `year`, `max`, `avg` FROM `score` limit 0, 100'
        cursor.execute(sql)
        rows = cursor.fetchall()

        #写入数据
        wb = Workbook()
        ws = wb.active
        for (i, row) in enumerate(rows):
            print(row)
            (ws['A{0}'.format(i + 1)],
            ws['B{0}'.format(i + 1)],
            ws['C{0}'.format(i + 1)]) = row

        # 保存excel
        wb.save('./static/export.xlsx')

if __name__ == "__main__":
    client = ExcelUtils()
    # client.do_sth()
    # client.read_xls()
    client.export_xls()
